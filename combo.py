import logging
import sys
from logging.handlers import RotatingFileHandler
import pandas as pd
from fuzzywuzzy import fuzz;
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
from nltk.tokenize import word_tokenize
import string
import requests
from requests.auth import HTTPBasicAuth
import json
from flask import Flask, render_template, request, redirect, url_for, session, flash
import nltk
from datetime import datetime, timedelta
import pytz
import re
import os
import time
import mimetypes
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from email_bot import send_email

parent_directory = os.path.dirname(os.path.dirname(os.path.dirname(os.path.realpath(__file__))))

nltk.download('punkt_tab')
nltk.download('punkt')
nltk.download('stopwords')
nltk.download('wordnet')
nltk.download('words')
words = set(nltk.corpus.words.words())


class ChatBot:
    def __init__(self, excel_file='input/details.xlsx'):
        self.responses = {}
        try:
            with open('input/responses.json') as f:
                self.responses = json.load(f)
        except FileNotFoundError:
            print("Error: responses.json not found.")
        self.lemmatizer = WordNetLemmatizer()
        self.df = pd.read_excel(excel_file)
        self.columns_to_search = ['Solution Name', 'Master Usecase Name', 'Solution Description']
        # Preprocess the loaded Excel data
        self.preprocess_excel_data()
        with open('config/config.json', 'r') as config_file:
            #logger.info("Json file read successfully")
            print("Json file read successfully")
            config_data = json.load(config_file)
            self.instance_url = config_data['instance_url']
            self.username = config_data['username']
            self.password = config_data['password']

    def is_gibberish(self, input_text):
        input_words = nltk.wordpunct_tokenize(input_text)
        if any(word.lower() not in words for word in input_words):
            return True
        return False

    def preprocess_text(self, text):
        text = str(text)
        text = text.lower()
        text = text.translate(str.maketrans('', '', string.punctuation))
        text = text.strip()
        stop_words = set(stopwords.words('english'))
        tokens = word_tokenize(text)
        filtered_tokens = [token for token in tokens if token not in stop_words]
        lemmatized_tokens = [self.lemmatizer.lemmatize(token) for token in filtered_tokens]
        preprocessed_text = ' '.join(lemmatized_tokens)
        # print(preprocessed_text)
        return preprocessed_text

    def get_response(self, user_input):
        return_respond_to_json = self.respond_to_json(user_input)
        if return_respond_to_json:
            return return_respond_to_json
        elif user_input.lower() and len(user_input) < 3:  # or self.is_gibberish(user_input)
            return "Please refine your search and try again"
        else:
            exact_match_response = self.get_exact_match_response(user_input)
            if exact_match_response:
                return exact_match_response
            partial_match_response = self.get_partial_match_response(user_input)
            if partial_match_response:
                return partial_match_response
            else:
                return "No match found. Please provide all details."

    def preprocess_excel_data(self):
        for column in self.columns_to_search:
            if column in self.df.columns:
                self.df[column] = self.df[column].apply(self.preprocess_text)

    def get_exact_match_response(self, user_input):
        user_input_processed = self.preprocess_text(user_input)
        for column in self.columns_to_search:
            if column in self.df.columns:
                # Using token sort ratio for unordered matching
                exact_result = self.df[self.df[column].apply(
                    lambda x: fuzz.token_sort_ratio(x, user_input_processed) == 100)]
                if not exact_result.empty:
                    responses = ["Exact match found: \n"
                                 "Instructions: 1.Verify the Solution name \n 2: Check the Functional Specification, to conclude which given solution is a close match with your Automation \n 3: Whichever given solution matches your requirement, paste the Id Number (format: 'ID XX') in the chat and click on submit button \n 4: If you want to create a ticket, click on create ticket"]
                    for _, row in exact_result.iterrows():
                        responses.extend([f'{col}: {row[col]}' for col in self.df.columns])
                    return '\n'.join(responses)

    def get_partial_match_response(self, user_input):
        user_input_processed = self.preprocess_text(user_input)
        matches = pd.DataFrame()
        for column in self.columns_to_search:
            if column in self.df.columns:
                processed_column = self.df[column].apply(self.preprocess_text)
                # Use token set ratio for more flexible matching
                result = self.df[processed_column.apply(
                    lambda x: fuzz.token_set_ratio(x, user_input_processed) > 63)]
                if not result.empty:
                    matches = pd.concat([matches, result])
        matches = matches.drop_duplicates()
        if not matches.empty:
            if len(matches) > 30:
                return 'Please refine your search and try again.'
            else:
                # Display a pop-up for partial match
                responses = ["Partial match found: \n"
                             "Instructions: 1.Verify the Solution name \n 2: Check the Functional Specification, to conclude which given solution is a close match with your Automation \n 3: Whichever given solution matches your requirement, paste the Id Number (format: 'ID XX') in the chat and click on submit button \n 4: If you want to create a ticket, click on create ticket"]
                responses.extend(
                    ['\n'.join(f'{k}: {v}' for k, v in match.to_dict().items()) for _, match in matches.iterrows()])
                return '\n---\n'.join(responses)

    def respond_to_json(self, user_input, ):
        user_input_lower = user_input.lower()
        best_match = self.find_best_match(user_input_lower)
        if best_match is not None:
            return self.responses[best_match]
        else:
            return None

    def find_best_match(self, user_input_lower):
        keys = list(self.responses.keys())
        best_match = None
        max_similarity = -1
        for key in keys:
            similarity = fuzz.token_sort_ratio(user_input_lower, key.lower())
            if similarity > max_similarity:
                max_similarity = similarity
                best_match = key
        return best_match if max_similarity >= 60 else None

    def safe_delete(self, file_path, max_attempts=3, delay=1):
        """Try to delete a file with a specified number of attempts and delay."""
        for attempt in range(max_attempts):
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                print(f"File {file_path} deleted successfully.")
                logging.info(f"File {file_path} deleted successfully.")
                break
            except PermissionError as e:
                print(f"Attempt {attempt + 1} failed: {e}")
                time.sleep(delay)
        else:
            print(f"Could not delete file {file_path} after {max_attempts} attempts.")

    def get_sys_id(self, instance, username, password, ticket_number):
        # ticket_number = response.json()['number']
        table_name = 'sc_request'
        query_url = f"{instance}/api/now/table/{table_name}?sysparm_query=number={ticket_number}&sysparm_fields=sys_id"
        # print(f"Query URL: {query_url}")  # Logging the query URL for debugging
        response = requests.get(query_url, auth=(username, password), verify=False)
        if response.status_code == 200:
            result = response.json().get('result')
            if result:
                return result[0].get('sys_id')
            else:
                # print("No record found.")
                return None
        else:
            print(f"Failed to retrieve sys_id: {response.status_code} {response.text}")
            logging.info(f"Failed to retrieve sys_id: {response.status_code} {response.text}")
            return None

    def attach_file_to_servicenow(self, instance, username, password, ticket_number, file_path):
        # ticket_number = ticket_number
        table_name = 'sc_request'
        record_sys_id = self.get_sys_id(instance, username, password, ticket_number)
        if record_sys_id is None:
            return
        file_name = os.path.basename(file_path)
        url = f'{instance}/api/now/attachment/file?table_name={table_name}&table_sys_id={record_sys_id}&file_name={file_name}'
        mime_type, _ = mimetypes.guess_type(file_path)
        if not mime_type:
            mime_type = 'application/octet-stream'
        with open(file_path, 'rb') as file:
            headers = {'Content-Type': mime_type}
            response = requests.post(url, auth=(username, password), headers=headers, data=file,verify=False)
            if response.status_code == 201:
                print("File attached successfully.")
                logging.info("File attached successfully.")
                return response.json()
            else:
                print(f"Failed to attach file: {response.status_code} {response.text}")
                return response.status_code, response.text

    def create_servicenow_ticket(self, data, short_description, special_instructions, file_paths=[],):
        table_api = 'sc_request'
        api_url = f'{self.instance_url}/api/now/table/{table_api}'
        headers = {
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        }

        # Create the ServiceNow ticket description
        description = ''
        for key, value in data.items():
            description += f"{key}: {value}\n"

        # Create the ServiceNow ticket
        payload = {
            'description': description,
            'short_description': short_description,
            'special_instructions': special_instructions
        }

        response = requests.post(api_url, auth=HTTPBasicAuth(self.username, self.password), headers=headers, json=payload, verify=False)
        if response.status_code != 201:
            return 'Failed to create ServiceNow ticket: ' + response.text
        ticket_number = response.json()['result']['number']

        # Create DataFrame from data
        df = pd.DataFrame(list(data.items()), columns=['Field', 'Value'])
        df.at[len(df), 'Field'] = 'Short Description'
        df.at[len(df)-1, 'Value'] = short_description
        df.at[len(df), 'Field'] = 'Special Instructions'
        df.at[len(df)-1, 'Value'] = special_instructions

        # Save form data to Excel file
        form_data_file_path = f"form_data-{ticket_number}.xlsx"
        wb = Workbook()
        ws = wb.active

        # Add headers to the worksheet
        headers = df.columns.tolist()
        ws.append(headers)

        # Add data to worksheet
        for index, row in df.iterrows():
            ws.append(row.tolist())

        # Add color to fields
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow color for headers
        for row in ws.iter_rows(min_row=2, max_row=len(df)+1):
            for cell in row:
                cell.fill = PatternFill(start_color="C4D9FF", end_color="C4D9FF", fill_type="solid")  # Green color for data cells

        wb.save(form_data_file_path)

        # Attach form data file to the ticket
        self.attach_file_to_servicenow(self.instance_url, self.username, self.password, ticket_number,
                                       file_path=form_data_file_path)
        self.safe_delete(file_path=form_data_file_path)

        # Attach additional files to the ticket
        for file_path in file_paths:
            self.attach_file_to_servicenow(self.instance_url, self.username, self.password, ticket_number,
                                           file_path=file_path)
            self.safe_delete(file_path=file_path)

        return 'ServiceNow ticket created successfully: ' + ticket_number


    def get_ticket_state(self, ticket_number):
        try:
            url = f'{self.instance_url}/api/now/table/sc_request?sysparm_query=number={ticket_number}&sysparm_fields=number,state&sysparm_limit=1'
            headers = {"Content-Type": "application/json", "Accept": "application/json"}
            response = requests.get(url, auth=(self.username, self.password), headers=headers, verify=False)
            response.raise_for_status()
            data = response.json()
            if 'result' in data and data['result']:
                ticket_state = data['result'][0]['state']
                print(
                    f'Ticket Number: {ticket_number}, Ticket State: {"In-progress" if ticket_state == "1" else "Closed"}')
                return f'Ticket Number: {ticket_number}, Ticket State: {"In-progress" if ticket_state == "1" else "Closed"}'
            else:
                print(f'No ticket found with number {ticket_number}')
                return f'No ticket found with number {ticket_number}'
        except requests.exceptions.HTTPError as errh:
            return f"HTTP Error: {errh}"
        except requests.exceptions.ConnectionError as errc:
            return f"Error Connecting: {errc}"
        except requests.exceptions.Timeout as errt:
            return f"Timeout Error: {errt}"
        except requests.exceptions.RequestException as err:
            return f"An unexpected error occurred: {err}"


app = Flask(__name__)
app.secret_key = 'Anuja@123'
chatbot = ChatBot()
# Set the session timeout to 10 minutes (600 seconds)
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=15)

############### Configure the logger ################
logs_directory = os.path.abspath(os.path.join(os.path.dirname(__file__), 'log'))
if not os.path.exists(logs_directory):
    os.makedirs(logs_directory)

log_file_path = os.path.join(logs_directory, "chatbot_logs.txt")
handler = RotatingFileHandler(log_file_path, maxBytes=10000, backupCount=1)
handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
logger.addHandler(handler)

#logger.info("Testing logging with FileHandler...")


# Before request function to check session timeout
@app.before_request
def check_session_timeout():
    if 'last_interaction' in session:
        last_interaction_time = session['last_interaction']
        current_time = datetime.now(pytz.utc)
        # Calculate the time difference
        time_difference = current_time - last_interaction_time
        # If the time difference exceeds the session timeout, clear the session
        if time_difference > app.permanent_session_lifetime:
            session.clear()
            flash("Your session has expired due to inactivity. Please start a new session.")
            return redirect(url_for('index'))
    # Update the last interaction time to the current time
    session['last_interaction'] = datetime.now(pytz.utc)


@app.route('/', methods=['GET', 'POST'])
def index():
    if 'new_session' not in session:
        flash(
            "Welcome to a smarter & seamless experience! Supercharge your efficiency with the CREATE Chatbot! Greet it to unleash powerful automation. Let the chatbot streamline tasks, saving time and eliminating errors.")
        session['new_session'] = False
        logger.info("Index page accessed")
    return render_template('index.html')



@app.route('/result', methods=['POST'])
def handle_form_submission():
    user_input = request.form.get('user_input')
    response = chatbot.get_response(user_input)
    # Check if the user input matches the pattern "id+any number" and there are search results, then redirect to the create_ticket_form
    match = re.match(r'id (\d+)', user_input.lower())
    if match:
        if 'search_results_found' in session and session['search_results_found']:
            id_value = match.group(1)  # Extracted ID value
            # Pass the ID as a query parameter
            return redirect(url_for('Options', id=id_value))
    else:
        # Check for search results and set the session variable accordingly
        response = chatbot.get_response(user_input)
    if 'Partial match found:' in response or 'Exact match found:' in response:
        session['search_results_found'] = True
    else:
        session['search_results_found'] = False
    # Check for "No match found" response
    if "No match found. Please provide all details." in response:
        if 'consecutive_attempts' in session:
            session['consecutive_attempts'] += 1
            if session['consecutive_attempts'] >= 2:
                response_data = [{'Leo': 'No match found Click on create ticket'}]
                session.pop('consecutive_attempts')
                return render_template('index.html', response_data=response_data)
        else:
            session['consecutive_attempts'] = 1
            flash("No match found. Please provide more details to search again.")
        return redirect(url_for('index'))
    else:
        if 'consecutive_attempts' in session:
            session.pop('consecutive_attempts')
        session.pop('_flashes', None)
        # Check if user input contains 'windows' or 'linux'
        if 'windows' == user_input.lower():
            return render_template('Options.html', platform='Windows')
        elif 'linux' == user_input.lower():
            return render_template('Options.html', platform='Linux')
        elif 'mssql' == user_input.lower():
            return render_template('Options.html', platform='MSSQL')
        elif 'oracle' == user_input.lower():
            return render_template('Options.html', platform='Oracle')
        elif 'sql' == user_input.lower():
            return render_template('Options.html', platform='sql')
        elif 'redhat' == user_input.lower():
            return render_template('Options.html', platform='Redhat')
        elif 'patching' == user_input.lower():
            return render_template('Options.html', platform='patching')
        elif 'health check' == user_input.lower():
            return render_template('Options.html', platform='health check')
        elif 'uptime report' == user_input.lower():
            return render_template('Options.html', platform='uptime report')
        elif 'log purging' == user_input.lower():
            return render_template('Options.html', platform='log purging')
        elif 'sccm' == user_input.lower():
            return render_template('Options.html', platform='sccm')
        else:
            # Your existing code for rendering the result
            response_data = []
            for entry in response.split('---'):
                entry_data = {}
                lines = entry.strip().split('\n')
                for line in lines:
                    if ': ' in line:
                        key, value = line.split(': ', 1)
                        entry_data[key] = value
                    else:
                        entry_data['Leo'] = line
                if entry_data:
                    response_data.append(entry_data)
            return render_template('index.html', response_data=response_data)


@app.route('/ticket_status', methods=['GET', 'POST'])
def ticket_status():
    if request.method == 'POST':
        ticket_number = request.form['ticket_number']
        return redirect(url_for('ticket_status_result', ticket_number=ticket_number))
    return render_template('ticket_status.html')


@app.route('/ticket_status_result/<ticket_number>', methods=['GET'])
def ticket_status_result(ticket_number):
    result = chatbot.get_ticket_state(ticket_number)
    print(f"Result from get_ticket_state: {result}")
    return render_template('ticket_status_result.html', result=result)


@app.route('/Options')
def Options():
    id_value = request.args.get('id')
    if id_value:
        session['id_value'] = id_value  # Storing id_value in the session
    return render_template('Options_result.html', id_value=id_value)


@app.route('/Options_result', methods=['GET', 'POST'])
def Options_result():
    if request.method == 'POST':
        platform = request.form.get('platform')
        request_type = request.form.get('request_type')
    elif request.method == 'GET':
        platform = request.args.get('platform')
        request_type = request.args.get('request_type')
    else:
        return "Method not allowed", 405
    # Ensure both platform and request_type have values before proceeding
    if not platform or not request_type:
        return "Invalid request", 400
    # Use a separator (e.g., '-') to combine platform and request_type
    user_input = f"{platform}-{request_type}"
    # Call the get_partial_match_response function
    response = chatbot.get_response(user_input)
    response_data = []
    if response:
        for entry in response.split('---'):
            entry_data = {}
            lines = entry.strip().split('\n')
            for line in lines:
                if ': ' in line:
                    key, value = line.split(': ', 1)
                    entry_data[key] = value
            if entry_data:
                response_data.append(entry_data)
    # If no data found, add a default entry
    if not response_data:
        response_data.append({
            'Leo': 'Oops! It seems that we do not have a matching result, kindly enter "create a ticket" & hit on submit button'})
    return render_template('index.html', response_data=response_data)


@app.route('/create_ticket_form', methods=['GET', 'POST'])
def create_ticket_form():
    id_value = session.get('id_value')
    if request.method == 'GET':
        if 'attempt' in session:
            session.pop('attempt')
            return render_template('create_ticket.html', id_value=id_value)
        return render_template('create_ticket.html', id_value=id_value)
    elif request.method == 'POST':
        description = request.form['description']
        short_description = request.form['short_description']
        special_instructions = request.form['special_instructions']
        ticket_response = chatbot.create_servicenow_ticket(description, short_description, special_instructions)
        session.pop('id_value', None)
        return render_template('ticket_result.html', ticket_response=ticket_response)


@app.route('/create_ticket', methods=['GET', 'POST'])
def create_ticket():
    if request.method == 'POST':
        # Extract form data
        short_description = request.form.get('short_description')
        userid = request.form.get('Bot_reference_ID')
        account_name = request.form.get('account_name')
        account_region = request.form.get('Account Region')
        account_spoc_details = request.form.get('account_spoc_details')
        ras_server = request.form.get('ras_server')
        whitelist_links = request.form.get('whitelist_links')
        test_environment = request.form.get('test_environment')
        authentication_required = request.form.get('authentication_required')
        api_call_required = request.form.get('api_call_required')
        python_version = request.form.get('select_language')
        special_instructions = request.form.get('special_instructions')

        # Prepare data for ticket creation
        data = {
            'Reference ID': userid,
            'Account Name': account_name,
            'Account Region': account_region,
            'Account SPOC Details': account_spoc_details,
            'Is JUMP/RAS server available and does it have Internet Connectivity ?': ras_server,
            'In case, Jump_internet connectivity, is it possible to whitelist few links for installing required modules?': whitelist_links,
            'If there is any authentication required, can a service account be arranged while testing & deployment ?': test_environment,
            'If API call to access details from any 3rd party software is needed, could API creds and instance be arranged while testing & deployment ? (Avaya API)': api_call_required,
            'If automation requirement is in Python, Kindly provide the version which is installed on the JUMP server or where the bot shall be triggered from. (e.g., Python 3.6, 3.8, 3.9, ...), write N/A otherwise.': python_version
        }

        # Use getlist to handle multiple files if necessary
        uploaded_files = request.files.getlist("myfile")
        file_paths = []
        # Process each uploaded file
        for file in uploaded_files:
            if file:
                filename = secure_filename(file.filename)
                file_path = os.path.join(parent_directory, filename)
                file.save(file_path)
                file_paths.append(file_path)

        # Call your chatbot or ServiceNow ticket creation function
        ticket_response = chatbot.create_servicenow_ticket(data, short_description, special_instructions, file_paths)
        # Optionally store ticket number or response in session for confirmation display
        session['ticket_response'] = ticket_response
        # Redirect to a confirmation page
        return redirect(url_for('ticket_result'))

    return render_template('create_ticket.html')
        #print(ticket_response)

@app.route('/ticket_result')
def ticket_result():
    ticket_response = session.get('ticket_response', 'Navigate back to the homepage to submit a new request')
    # Clear ticket_response from session after retrieving it
    session.pop('ticket_response', None)
    return render_template('ticket_result.html', ticket_response=ticket_response)

@app.route('/Redeployemt_ticket_form', methods=['GET', 'POST'])
def Redeployemt_ticket_form():
    id_value = session.get('id_value')
    if request.method == 'GET':
        if 'search_results_found' in session:
            session.pop('search_results_found')
            return render_template('Redeployemt.html', id_value=id_value)
        return render_template('Redeployemt.html')
    elif request.method == 'POST':
        description = request.form['description']
        short_description = request.form['short_description']
        special_instructions = request.form['special_instructions']
        ticket_response = chatbot.create_servicenow_ticket(description, short_description, special_instructions)
        session.pop('id_value', None)
        return render_template('ticket_result.html', ticket_response=ticket_response)


@app.route('/redeployemt_ticket', methods=['POST'])
def redeployemt_ticket():
    id_value = session.get('id_value')
    if request.method == 'POST':
        # Retrieve form data using individual field names
        demand_owner_name = request.form.get('1.Demand Owner/Automation SPOC Name')
        demand_owner_email = request.form.get('2.Demand Owner/Automation SPOC E-mail ID')
        account_region = request.form.get("Account's Region")
        account_sub_region = request.form.get("Account's Sub Region")
        uk_account = request.form.get('5.Account - For The UK:')
        # Concatenate the values if needed
        data = {
            'Demand Owner/Automation SPOC Name': demand_owner_name,
            'Demand Owner/Automation SPOC E-mail ID': demand_owner_email,
            "Account's Region": account_region,
            "Account's Sub Region": account_sub_region,
            'Select Account Name': uk_account,
        }
        description = f"Demand Owner/Automation SPOC Name: {demand_owner_name}\n" \
                      f"Demand Owner/Automation SPOC E-mail ID: {demand_owner_email}\n" \
                      f"Account's Region: {account_region}\n" \
                      f"Account's Sub Region: {account_sub_region}\n" \
                      f"Select Account Name: {uk_account}"
        short_description = request.form.get('short_description')
        special_instructions = request.form.get('special_instructions')
        # Use getlist to handle multiple files if necessary
        uploaded_files = request.files.getlist("myfile")
        file_paths = []
        # Process each uploaded file
        for file in uploaded_files:
            if file:
                filename = secure_filename(file.filename)
                file_path = os.path.join(parent_directory, filename)
                file.save(file_path)
                file_paths.append(file_path)
        # Pass the file paths to your ticket creation method
        ticket_response = chatbot.create_servicenow_ticket(data, short_description, special_instructions, file_paths)
        # Send email using ticket number and demand owner email
        # send_email(ticket_response, demand_owner_email)
        print(ticket_response)

        session.pop('id_value', None)
        # Store ticket response in session for use on the result page
        session['ticket_response'] = ticket_response

        # Redirect to the result page
        return redirect(url_for('ticket_result'))

    # In case of GET requests or other methods, render the form
    return render_template('Redeployemt.html')

if __name__ == '__main__':
    app.run(host='0.0.0.0',debug=True)
